# -*- coding: utf-8 -*-
"""
원본 엑셀(DS 설치현황 시트) → webapp seed.json / seed.js  (정합성 강화 v2)

핵심 수정:
 - 숫자는 엑셀 표시서식(number_format)대로 포맷 ('0'→정수, 'General'→실제값)
 - 열수(row4)는 병합셀 → 각 열로 전파
 - SIZE(row44) "W\nH" → "W×H"
 - 상태는 레이어 문맥 반영(바닥=타공, 입상/횡주=설치)
 - 대각선 = 횡주간 없음(횡주 레이어 전용, 값 없음)
 - 상부접점(row5)/하부접점(row36) 구조행 포함
 - 엑셀 자체 타공현황 카운트(row35) 보존
사용법: python import_xlsx.py
"""
import openpyxl, json, re, os

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_NAME = "(260617)세보 PH2 DUCT설치현황_rev01.xlsx"
# 저장소 루트(= webapp 상위)에 둔 원본을 기본 사용. 환경변수 DS_XLSX 로 덮어쓰기 가능.
SRC = os.environ.get("DS_XLSX") or os.path.join(HERE, "..", XLSX_NAME)
SHEET = "DS 설치현황"
OUT = os.path.join(HERE, "..", "seed.json")   # 배포 폴더(webapp) 밖 — GitHub Pages 에 올라가지 않음(앱은 seed.js만 사용)

THEME = ['FFFFFF','000000','E7E6E6','44546A','5B9BD5','ED7D31','A5A5A5','FFC000','4472C4','70AD47','0563C1','954F72']

def apply_tint(h, t):
    def f(c): return c*(1+t) if t < 0 else c*(1-t) + 255*t
    return '%02X%02X%02X' % (round(f(int(h[0:2],16))), round(f(int(h[2:4],16))), round(f(int(h[4:6],16))))

def cell_rgb(c):
    f = c.fill
    if not f or f.patternType is None: return None
    fg = f.fgColor
    if fg is None: return None
    if fg.type == 'rgb' and fg.rgb and fg.rgb != '00000000': return fg.rgb[-6:].upper()
    if fg.type == 'theme':
        base = THEME[fg.theme] if fg.theme < len(THEME) else 'FFFFFF'
        return apply_tint(base, fg.tint or 0.0)
    return None

def fmt_num(v, nf):
    """물량은 항상 정수로 반올림 표기(사용자 요청)."""
    if not isinstance(v, (int, float)): return None
    return str(int(round(v)))

# 색 → 상태  (범례 이미지 직접 확인: 색 스와치는 라벨 '왼쪽'에 배치)
GRAY = {'BFBFBF', 'BEBEBE', 'C0C0C0'}
NONE_C = {None, 'FFFFFF'}
HEADER_BEIGE = 'FBE5D7'

# 금일색(노랑·진파랑)은 '완료 + 당일' 의 일시적 표현 → 저장은 완료상태로,
# 당일 여부는 날짜(d)로 렌더 시 판정. (다음날 자동 완료색)
def resolve_status(rgb, layer):
    if rgb == HEADER_BEIGE or rgb in NONE_C: return 'none'            # 흰색 = 해당없음
    if rgb in GRAY: return 'not_installed'                            # 회색 = 미설치
    if rgb == 'FF0000': return 'etc_interf'                           # 빨강 = 기타 간섭구간
    if rgb == 'FF8F8F': return 'scaffold_interf'                      # 핑크 = 비계 간섭구간
    if rgb == 'FFFF00': return 'drill_done'                           # 노랑 = 금일타공 → 타공완료(+당일)
    if rgb == '0070C0': return 'install_done'                         # 진파랑 = 금일설치 → 설치완료(+당일)
    if rgb == '66FFFF': return 'drill_done' if layer == '바닥' else 'install_done'      # 시안 = 타공완료/설치완료
    if rgb == '00B0F0': return 'predrill_floor' if layer == '바닥' else 'predrill_duct'  # 하늘 = 기설치(타공/덕트)
    return 'none'

TODAY_COLORS = {'FFFF00', '0070C0'}   # 원본에서 '금일'로 칠해져 있던 색

ZONES = [
    ('북DS(FA,SA)', 3, 22),
    ('동DS(FA)',    23, 25),
    ('북DS(배기)',  27, 98),
    ('동DS(배기)',  100, 104),
]
FLOOR_ORDER = ['10F','9F','8F','7F','6F','5F','4F','3F','2F','1F']

def parse_row_layer(label):
    if label is None: return None
    s = str(label).strip().replace(' ', '')
    if re.match(r'^(\d+)[Ff]?횡주', s): return (re.match(r'^(\d+)', s).group(1)+'F', '횡주')
    if re.match(r'^(\d+)[Ff층충]+바닥', s): return (re.match(r'^(\d+)', s).group(1)+'F', '바닥')
    if re.match(r'^(\d+)[Ff]$', s): return (re.match(r'^(\d+)', s).group(1)+'F', '입상')
    return None

def main():
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]

    # 행 매핑
    row_map = {}
    for r in range(5, 36):
        fl = parse_row_layer(ws.cell(row=r, column=1).value)
        if fl: row_map[r] = fl

    # 열수(row4) 병합 전파
    yeolsu = {}
    for m in ws.merged_cells.ranges:
        if m.min_row <= 4 <= m.max_row:
            v = ws.cell(row=4, column=m.min_col).value
            if v and '열' in str(v):
                for col in range(m.min_col, m.max_col+1): yeolsu[col] = str(v).strip()
    for col in range(3, 105):
        if col in yeolsu: continue
        v = ws.cell(row=4, column=col).value
        if v and '열' in str(v): yeolsu[col] = str(v).strip()

    def gv(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None: return None
        if hasattr(v, 'isoformat'): return v.strftime('%Y-%m-%d')
        return v
    def size_fmt(v):
        if v is None: return None
        return str(v).replace('\n', '×').strip()

    parts, cells = [], []
    for zname, c0, c1 in ZONES:
        for col in range(c0, c1+1):
            part_no = gv(37, col)
            if part_no is None or str(part_no).strip() == '': continue
            L = get_column_letter(col)
            seong = gv(38, col); seong = str(seong).strip().replace('\n','') if seong is not None else None
            parts.append({
                'id': L, 'col': col, 'col_letter': L, 'zone': zname,
                'part_no': str(part_no).strip(),
                'seong': seong,
                'size': size_fmt(gv(44, col)),
                'yeolsu': yeolsu.get(col),
                'to': (str(gv(39, col)).replace('\n',' ').strip() if gv(39, col) else None),
                'takong_excel': gv(35, col),   # 엑셀 자체 타공현황 카운트
                'upper': None, 'lower': None,   # 상/하부접점(현재 빈값, 편집 가능)
            })
            for r, (floor, layer) in row_map.items():
                c = ws.cell(row=r, column=col)
                rgb = cell_rgb(c)
                b = c.border
                diag = bool(b and b.diagonal and b.diagonal.style)
                raw = c.value
                if isinstance(raw, str) and raw.startswith('='): raw = None
                qty = raw if isinstance(raw, (int, float)) else None
                qty_disp = fmt_num(raw, c.number_format) if isinstance(raw, (int, float)) else None
                # 상태
                if layer == '횡주' and diag:
                    status = 'no_beam'           # 횡주간 없음
                else:
                    status = resolve_status(rgb, layer)
                # 원본에서 '금일'(노랑/진파랑)로 칠해진 칸은 기준일(updated)에 완료된 것으로 기록
                d = (str(ws['A1'].value)[:10] if ws['A1'].value else None) if (rgb in TODAY_COLORS) else None
                cells.append({
                    'part': L, 'zone': zname, 'floor': floor, 'layer': layer,
                    'status': status,
                    'd': d,
                    'qty': (round(qty, 3) if isinstance(qty, float) else qty),
                    'qd': qty_disp,
                    'diag': diag,
                    'ref': f"{L}{r}",
                })

    seed = {
        'sheet': SHEET, 'generated_from': os.path.basename(SRC),
        'updated': str(ws['A1'].value)[:10] if ws['A1'].value else None,
        'zones': [z[0] for z in ZONES],
        'zone_ranges': {z[0]: [z[1], z[2]] for z in ZONES},
        'floors': FLOOR_ORDER,
        'layers': ['횡주','입상','바닥'],
        # sel=true 만 편집 팔레트에 노출. 금일색은 auto(완료+당일 자동).
        'legend': [
            {'key':'not_installed','label':'미설치','color':'#BFBFBF','layers':['횡주','입상','바닥'],'sel':True},
            {'key':'install_done','label':'설치완료','color':'#66FFFF','layers':['횡주','입상'],'sel':True},
            {'key':'drill_done','label':'타공완료','color':'#66FFFF','layers':['바닥'],'sel':True},
            {'key':'etc_interf','label':'기타 간섭구간','color':'#FF0000','layers':['횡주','입상','바닥'],'sel':True},
            {'key':'scaffold_interf','label':'비계 간섭구간','color':'#FF8F8F','layers':['횡주','입상','바닥'],'sel':True},
            {'key':'predrill_duct','label':'기설치덕트','color':'#00B0F0','layers':['횡주','입상'],'sel':True},
            {'key':'predrill_floor','label':'기설치타공','color':'#00B0F0','layers':['바닥'],'sel':True},
            {'key':'today_install','label':'금일설치','color':'#0070C0','layers':['횡주','입상'],'sel':False,'auto':True},
            {'key':'today_drill','label':'금일타공','color':'#FFFF00','layers':['바닥'],'sel':False,'auto':True},
            {'key':'none','label':'해당없음','color':'#FFFFFF','layers':['횡주','입상','바닥'],'sel':False},
            {'key':'no_beam','label':'횡주간 없음','color':'#EAEAEA','layers':['횡주'],'sel':False},
        ],
        'parts': parts, 'cells': cells,
    }
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(seed, f, ensure_ascii=False, indent=1)
    with open(os.path.join(HERE, 'seed.js'), 'w', encoding='utf-8') as f:   # seed.js 는 앱이 로드 → webapp 안에 유지
        f.write('window.__SEED__ = '); json.dump(seed, f, ensure_ascii=False); f.write(';\n')

    from collections import Counter
    print("parts=%d cells=%d" % (len(parts), len(cells)))
    print("status:", dict(Counter(c['status'] for c in cells)))
    print("열수 채움:", sum(1 for p in parts if p['yeolsu']), "/", len(parts))
    print("SIZE 채움:", sum(1 for p in parts if p['size']), "/", len(parts))
    # 검증용: 9F 입상(row10) 배기 앞부분 표시값
    sample = [c['qd'] for c in cells if c['floor']=='9F' and c['layer']=='입상' and c['zone']=='북DS(배기)'][:14]
    print("9F 입상 배기 표시값:", sample)

if __name__ == '__main__':
    main()
